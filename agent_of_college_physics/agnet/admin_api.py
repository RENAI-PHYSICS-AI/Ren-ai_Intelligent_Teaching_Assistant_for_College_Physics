import os
import json
import hmac
import io
import re
import time
import tomllib
import zipfile
from pathlib import Path

from fastapi import Cookie, FastAPI, Header, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse
from pydantic import BaseModel, Field
from openpyxl import load_workbook

import analytics_db as db
import admin_auth
from proxy_paths import with_public_prefix
import user_session


BASE_DIR = Path(__file__).resolve().parent
SECRETS_PATH = BASE_DIR / ".streamlit" / "secrets.toml"
_AUTH_FAILURES: dict[str, list[float]] = {}
_AUTH_WINDOW_SECONDS = 300
_AUTH_FAILURE_LIMIT = 10
_ADMIN_SESSION_COOKIE = "physics_admin_session"
_ADMIN_SESSION_SECONDS = 8 * 60 * 60
_USED_LOGIN_NONCES: dict[str, int] = {}
_USED_USER_LOGIN_NONCES: dict[str, int] = {}
_MAX_EXCEL_BYTES = 8 * 1024 * 1024
_MAX_EXCEL_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
_MAX_ROSTER_ROWS = 2000


class IdentityRosterEntry(BaseModel):
    identity_type: str = Field(min_length=1, max_length=16)
    institutional_id: str = Field(min_length=1, max_length=64)
    real_name: str = Field(min_length=1, max_length=64)


class IdentityRosterBatch(BaseModel):
    entries: list[IdentityRosterEntry] = Field(min_length=1, max_length=2000)


class IdentityRosterEdit(IdentityRosterEntry):
    pass


def _normalize_excel_header(value) -> str:
    return re.sub(r"[\s_\-/（）()]+", "", str(value or "").strip().lower())


def _excel_cell_text(value) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_identity_roster_excel(content: bytes) -> tuple[list[dict], list[str]]:
    if not content:
        raise ValueError("Excel 文件为空。")
    if len(content) > _MAX_EXCEL_BYTES:
        raise ValueError("Excel 文件不能超过 8 MB。")

    source = io.BytesIO(content)
    if not zipfile.is_zipfile(source):
        raise ValueError("文件不是有效的 .xlsx 或 .xlsm 工作簿。")
    source.seek(0)
    with zipfile.ZipFile(source) as archive:
        members = archive.infolist()
        if len(members) > 2048 or sum(item.file_size for item in members) > _MAX_EXCEL_UNCOMPRESSED_BYTES:
            raise ValueError("Excel 文件解压后体积过大。")
        if any(item.flag_bits & 0x1 for item in members):
            raise ValueError("不支持加密的 Excel 文件。")

    source.seek(0)
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("无法读取 Excel 文件，请确认文件格式完整。") from exc

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_values = None
        header_row_number = 0
        for row_number, row in enumerate(rows, start=1):
            if any(value not in (None, "") for value in row):
                header_values = row
                header_row_number = row_number
                break
            if row_number >= 10:
                break
        if not header_values:
            raise ValueError("Excel 中没有表头。")

        aliases = {
            "identity_type": {"身份类型", "身份", "类型", "用户类型", "identitytype"},
            "institutional_id": {"学号工号", "学号", "工号", "编号", "institutionalid", "id"},
            "real_name": {"姓名", "真实姓名", "名字", "realname", "name"},
        }
        normalized_headers = [_normalize_excel_header(value) for value in header_values]
        indexes = {}
        for field, field_aliases in aliases.items():
            indexes[field] = next(
                (index for index, header in enumerate(normalized_headers) if header in field_aliases),
                None,
            )
        missing = [field for field, index in indexes.items() if index is None]
        if missing:
            raise ValueError("Excel 表头必须包含：身份类型、学号/工号、姓名。")

        entries = []
        errors = []
        for row_number, row in enumerate(rows, start=header_row_number + 1):
            values = list(row)
            if not any(value not in (None, "") for value in values):
                continue
            if len(entries) + len(errors) >= _MAX_ROSTER_ROWS:
                errors.append(f"超过 {_MAX_ROSTER_ROWS} 条的记录未导入")
                break

            def value_for(field):
                index = indexes[field]
                return _excel_cell_text(values[index] if index < len(values) else None)

            type_text = value_for("identity_type").lower()
            if type_text in {"学生", "student"}:
                identity_type = "student"
            elif type_text in {"教师", "老师", "teacher"}:
                identity_type = "teacher"
            else:
                errors.append(f"第 {row_number} 行身份类型无效")
                continue
            institutional_id = value_for("institutional_id")
            real_name = value_for("real_name")
            if not institutional_id or not real_name:
                errors.append(f"第 {row_number} 行学号/工号或姓名为空")
                continue
            entries.append({
                "identity_type": identity_type,
                "institutional_id": institutional_id,
                "real_name": real_name,
            })
        return entries, errors
    finally:
        workbook.close()


def _load_admin_token() -> str:
    env_token = (os.environ.get("ADMIN_TOKEN", "") or os.environ.get("ADMIN_ANALYTICS_TOKEN", "")).strip()
    if env_token:
        return env_token
    if SECRETS_PATH.exists():
        with SECRETS_PATH.open("rb") as f:
            secrets = tomllib.load(f)
        configured = str(secrets.get("admin_token", "")).strip()
        if configured:
            return configured
    return admin_auth.load_or_create_local_secret(BASE_DIR / "data" / "admin_signing_secret")


def _load_admin_username() -> str:
    env_username = os.environ.get("ADMIN_USERNAME", "").strip()
    if env_username:
        return env_username
    if SECRETS_PATH.exists():
        with SECRETS_PATH.open("rb") as f:
            secrets = tomllib.load(f)
        return str(secrets.get("admin_username", "admin")).strip() or "admin"
    return "admin"


def _load_admin_password() -> str:
    env_password = os.environ.get("ADMIN_PASSWORD", "")
    if env_password:
        return env_password
    if SECRETS_PATH.exists():
        with SECRETS_PATH.open("rb") as f:
            secrets = tomllib.load(f)
        return str(secrets.get("admin_password", ""))
    return ""


def _ensure_configured_admin() -> None:
    password = _load_admin_password()
    if len(password) < 12:
        return
    display_name = os.environ.get("ADMIN_DISPLAY_NAME", "").strip()
    if not display_name and SECRETS_PATH.exists():
        with SECRETS_PATH.open("rb") as f:
            display_name = str(tomllib.load(f).get("admin_display_name", "课程管理员")).strip()
    db.ensure_admin_user(_load_admin_username(), password, display_name or "课程管理员")


db.init_db()
_ensure_configured_admin()


def _valid_admin_session(token: str | None) -> bool:
    payload = admin_auth.verify_token(_load_admin_token(), token or "", "admin-session")
    if not payload:
        return False
    user = db.get_user_by_username(str(payload.get("sub", "")))
    return bool(user and user.get("role") == "admin" and user.get("is_active"))


def _require_admin_token(
    x_admin_token: str | None,
    authorization: str | None,
    client_ip: str = "unknown",
    admin_session: str | None = None,
) -> None:
    if _valid_admin_session(admin_session):
        return
    now = time.monotonic()
    recent_failures = [t for t in _AUTH_FAILURES.get(client_ip, []) if now - t < _AUTH_WINDOW_SECONDS]
    if len(recent_failures) >= _AUTH_FAILURE_LIMIT:
        raise HTTPException(status_code=429, detail="Too many authentication attempts.")

    expected = _load_admin_token()
    if not expected:
        raise HTTPException(status_code=503, detail="Admin token is not configured.")

    supplied = x_admin_token or ""
    if not supplied and authorization:
        scheme, _, token = authorization.partition(" ")
        if scheme.lower() == "bearer":
            supplied = token

    if not hmac.compare_digest(supplied, expected):
        recent_failures.append(now)
        _AUTH_FAILURES[client_ip] = recent_failures
        raise HTTPException(status_code=401, detail="Invalid admin token.")

    _AUTH_FAILURES.pop(client_ip, None)


def _analytics_payload(recent_error_limit: int = 15) -> dict:
    return {
        "total": db.get_total_stats(),
        "chapters": db.get_chapter_stats(),
        "daily": db.get_daily_stats(),
        "feedback": db.get_feedback_stats(),
        "feedback_recent": db.get_recent_feedback(),
        "users": db.get_user_stats(),
        "learning": db.get_learning_analytics(),
        "roster": db.get_identity_roster_stats(),
        "errors": {
            "by_type": db.get_error_stats(),
            "recent": db.get_recent_errors(recent_error_limit),
        },
        "unanswered": db.get_unanswered_questions(),
        "response_timings": db.get_recent_response_timings(),
    }


def _analytics_login_page(auto_load: bool = False, public_prefix: str = "") -> str:
    page = """
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>数据分析 - 管理员访问</title>
  <style>
    body { margin: 0; font-family: "Microsoft YaHei", system-ui, sans-serif; background: #0f141b; color: #e7edf5; }
    main { max-width: 1120px; margin: 0 auto; padding: 36px 20px; }
    .hero { background: #142235; border: 1px solid #26384f; border-radius: 10px; padding: 24px; margin-bottom: 22px; }
    h1 { margin: 0 0 8px; font-size: 28px; }
    .muted { color: #9fb0c4; }
    .login { display: flex; gap: 10px; align-items: center; margin-top: 18px; }
    input { flex: 1; min-width: 240px; padding: 11px 12px; border-radius: 6px; border: 1px solid #40546d; background: #0b1118; color: #fff; font-size: 15px; }
    button { padding: 11px 16px; border-radius: 6px; border: 0; background: #2f80ed; color: #fff; font-weight: 700; cursor: pointer; }
    button:hover { background: #4a90f0; }
    .error { color: #ff8a8a; margin-top: 12px; min-height: 22px; }
    .grid { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin: 18px 0; }
    .card { background: #121b27; border: 1px solid #26384f; border-radius: 8px; padding: 16px; }
    .metric { font-size: 26px; font-weight: 800; margin-top: 8px; }
    .cols { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
    .section-title { margin: 28px 0 6px; font-size: 24px; }
    .section-note { margin: 0 0 14px; color: #9fb0c4; }
    .table-wrap { overflow: auto; max-height: 520px; border: 1px solid #26384f; border-radius: 8px; }
    table { width: 100%; border-collapse: collapse; min-width: 920px; background: #121b27; }
    th, td { padding: 11px 12px; border-bottom: 1px solid #26384f; text-align: left; vertical-align: top; }
    th { position: sticky; top: 0; z-index: 1; background: #182638; color: #c8d7e8; }
    tr:last-child td { border-bottom: 0; }
    .status { display: inline-block; padding: 3px 8px; border-radius: 12px; background: #173d31; color: #8ce0b8; font-size: 12px; }
    .status.attention { background: #4a2d1c; color: #ffc184; }
    .bar { height: 8px; margin-top: 7px; overflow: hidden; border-radius: 4px; background: #26384f; }
    .bar > span { display: block; height: 100%; background: #4a90f0; }
    .rules { margin: 10px 0 0; padding-left: 20px; color: #9fb0c4; font-size: 13px; }
    .roster-form { display: grid; grid-template-columns: minmax(0, 1fr) auto; gap: 12px; align-items: end; }
    .roster-actions { display: inline-flex; gap: 6px; margin-left: 10px; vertical-align: middle; }
    .roster-actions button { padding: 4px 9px; font-size: 12px; }
    .roster-actions button.danger { background: #a43d4a; }
    .roster-actions button.danger:hover { background: #c4515d; }
    .excel-upload { display: grid; grid-template-columns: minmax(0, 1fr) auto; gap: 12px; align-items: center; margin-top: 14px; padding-top: 14px; border-top: 1px solid #26384f; }
    .excel-upload input[type="file"] { min-width: 0; width: 100%; box-sizing: border-box; }
    textarea { width: 100%; min-height: 130px; box-sizing: border-box; resize: vertical; padding: 11px 12px; border-radius: 6px; border: 1px solid #40546d; background: #0b1118; color: #fff; font: 14px/1.6 Consolas, monospace; }
    .success { color: #8ce0b8; min-height: 22px; margin-top: 8px; }
    pre { white-space: pre-wrap; word-break: break-word; background: #091019; border: 1px solid #26384f; border-radius: 8px; padding: 12px; max-height: 360px; overflow: auto; }
    li { margin: 6px 0; }
    @media (max-width: 800px) { .grid, .cols, .roster-form, .excel-upload { grid-template-columns: 1fr; } .login { flex-direction: column; align-items: stretch; } }
  </style>
</head>
<body>
  <main>
    <section class="hero">
      <h1>📈 大学物理智能助教数据分析</h1>
      <div class="muted">__AUTH_NOTE__</div>
      <div class="login" __AUTH_STYLE__>
        <input id="token" type="password" placeholder="输入管理员口令" autocomplete="current-password" />
        <button onclick="loadAnalytics()">进入</button>
      </div>
      <div id="error" class="error"></div>
    </section>
    <section id="dashboard"></section>
  </main>
  <script>
    const tokenInput = document.getElementById("token");
    tokenInput.addEventListener("keydown", event => {
      if (event.key === "Enter") loadAnalytics();
    });

    function itemList(rows, mapper) {
      if (!rows || rows.length === 0) return "<p class='muted'>暂无数据</p>";
      return "<ul>" + rows.map(row => mapper(new Proxy(row, {
        get(target, key) { return escapeHtml(target[key]); }
      }))).join("") + "</ul>";
    }

    function escapeHtml(value) {
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }

    function formatDuration(value) {
      const ms = Number(value);
      if (!Number.isFinite(ms)) return "—";
      return ms < 1000 ? `${ms.toFixed(1)} ms` : `${(ms / 1000).toFixed(ms >= 10000 ? 1 : 3)} s`;
    }

    async function importRoster() {
      const source = document.getElementById("roster-input");
      const resultEl = document.getElementById("roster-result");
      if (!source || !resultEl) return;
      resultEl.textContent = "";
      const entries = [];
      const errors = [];
      source.value.split(/\\r?\\n/).forEach((line, index) => {
        if (!line.trim()) return;
        const parts = line.split(/[,，\\t]/).map(value => value.trim());
        const typeText = (parts[0] || "").toLowerCase();
        const identityType = ["学生", "student"].includes(typeText) ? "student"
          : (["教师", "老师", "teacher"].includes(typeText) ? "teacher" : "");
        if (!identityType || !parts[1] || !parts[2]) {
          errors.push(`第 ${index + 1} 行格式不正确`);
          return;
        }
        entries.push({ identity_type: identityType, institutional_id: parts[1], real_name: parts[2] });
      });
      if (errors.length || entries.length === 0) {
        resultEl.textContent = errors.join("；") || "请输入至少一条名册记录。";
        return;
      }
      const resp = await fetch("/identity-roster", {
        method: "POST",
        headers: { "X-Admin-Token": tokenInput.value.trim(), "Content-Type": "application/json" },
        body: JSON.stringify({ entries })
      });
      const result = await resp.json();
      if (!resp.ok) {
        resultEl.textContent = result.detail || "名册导入失败。";
        return;
      }
      resultEl.textContent = `新增 ${result.added} 条，更新 ${result.updated} 条，未变化 ${result.unchanged} 条。${(result.errors || []).join("；")}`;
      source.value = "";
      setTimeout(loadAnalytics, 700);
    }

    async function editRoster(id, identityType, institutionalId, realName) {
      const nextId = prompt("编号（学号/工号）", institutionalId);
      if (nextId === null) return;
      const nextName = prompt("姓名", realName);
      if (nextName === null) return;
      const nextTypeText = prompt("身份类型：student 或 teacher", identityType);
      if (nextTypeText === null) return;
      const nextType = nextTypeText.trim().toLowerCase();
      if (!nextId.trim() || !nextName.trim() || !["student", "teacher"].includes(nextType)) {
        alert("身份类型、编号和姓名不能为空，类型只能是 student 或 teacher。");
        return;
      }
      const resp = await fetch(apiUrl(`/identity-roster/${id}`), {
        method: "PUT",
        headers: {"X-Admin-Token": tokenInput.value.trim(), "Content-Type": "application/json"},
        body: JSON.stringify({identity_type: nextType, institutional_id: nextId.trim(), real_name: nextName.trim()})
      });
      if (!resp.ok) { const result = await resp.json(); alert(result.detail || "修改失败"); return; }
      loadAnalytics();
    }

    async function deleteRoster(id, label) {
      if (!confirm(`确定删除名册记录“${label}”吗？`)) return;
      const resp = await fetch(apiUrl(`/identity-roster/${id}`), {
        method: "DELETE",
        headers: {"X-Admin-Token": tokenInput.value.trim()}
      });
      if (!resp.ok) { const result = await resp.json(); alert(result.detail || "删除失败"); return; }
      loadAnalytics();
    }

    async function importRosterExcel() {
      const fileInput = document.getElementById("roster-excel");
      const resultEl = document.getElementById("roster-result");
      const button = document.getElementById("roster-excel-button");
      if (!fileInput || !resultEl || !button) return;
      const file = fileInput.files && fileInput.files[0];
      if (!file) {
        resultEl.textContent = "请选择 Excel 文件。";
        return;
      }
      if (file.size > 8 * 1024 * 1024) {
        resultEl.textContent = "Excel 文件不能超过 8 MB。";
        return;
      }
      resultEl.textContent = "正在读取并导入 Excel 名册...";
      button.disabled = true;
      try {
        const resp = await fetch("/identity-roster/excel", {
          method: "POST",
          headers: {
            "X-Admin-Token": tokenInput.value.trim(),
            "Content-Type": "application/octet-stream"
          },
          credentials: "same-origin",
          body: file
        });
        const result = await resp.json();
        if (!resp.ok) {
          resultEl.textContent = result.detail || "Excel 名册导入失败。";
          return;
        }
        resultEl.textContent = `Excel 共读取 ${result.rows_read || 0} 条有效记录；新增 ${result.added} 条，更新 ${result.updated} 条，未变化 ${result.unchanged} 条。${(result.errors || []).join("；")}`;
        fileInput.value = "";
        setTimeout(loadAnalytics, 700);
      } catch (error) {
        resultEl.textContent = `Excel 名册导入失败：${error.message || error}`;
      } finally {
        button.disabled = false;
      }
    }

    async function loadAnalytics() {
      const token = tokenInput.value.trim();
      const errorEl = document.getElementById("error");
      const dashboard = document.getElementById("dashboard");
      errorEl.textContent = "";
      dashboard.innerHTML = "";
      const headers = token ? { "X-Admin-Token": token } : {};
      const resp = await fetch("/analytics?format=json", {
        headers,
        credentials: "same-origin"
      });
      if (!resp.ok) {
        const text = await resp.text();
        errorEl.textContent = resp.status === 401 ? "口令错误。" : text;
        return;
      }
      const data = await resp.json();
      const total = data.total || {};
      const users = data.users || {};
      const roster = data.roster || {};
      const learning = data.learning || {};
      const learningOverview = learning.overview || {};
      const learners = learning.learners || [];
      const chapterLearning = learning.chapters || [];
      const maxChapterQuestions = Math.max(1, ...chapterLearning.map(row => Number(row.question_count || 0)));
      const learnerRows = learners.length ? learners.map(row => {
        const attention = row.status === "attention";
        const lastActivity = row.last_activity ? String(row.last_activity).slice(0, 19) : "暂无";
        const errorRate = Number(row.error_rate || 0).toFixed(1);
        return `<tr>
          <td><b>${escapeHtml(row.display_name || row.username)}</b><br><span class="muted">@${escapeHtml(row.username)}</span></td>
          <td>${Number(row.question_count || 0)}</td>
          <td>${Number(row.learning_days || 0)}</td>
          <td>${Number(row.chapters_covered || 0)}</td>
          <td>${Number(row.session_count || 0)}</td>
          <td>${errorRate}%</td>
          <td>${Number(row.good_feedback || 0)} / ${Number(row.bad_feedback || 0)}</td>
          <td>${escapeHtml(lastActivity)}</td>
          <td><span class="status ${attention ? "attention" : ""}">${attention ? escapeHtml(row.attention_reason) : "学习中"}</span></td>
        </tr>`;
      }).join("") : `<tr><td colspan="9" class="muted">暂无注册用户学情数据</td></tr>`;
      const chapterLearningList = chapterLearning.length ? "<ul>" + chapterLearning.map(row => {
        const width = Math.max(3, Math.round(Number(row.question_count || 0) * 100 / maxChapterQuestions));
        return `<li><b>${escapeHtml(row.chapter)}</b>：${Number(row.question_count || 0)} 题，${Number(row.learner_count || 0)} 人，问答请求失败率 ${Number(row.error_rate || 0).toFixed(1)}%<div class="bar"><span style="width:${width}%"></span></div></li>`;
      }).join("") + "</ul>" : "<p class='muted'>暂无数据</p>";
      const dailyLearningList = itemList(learning.daily, row => `<li><b>${row.day}</b>：${row.learner_count || 0} 人，${row.question_count || 0} 题，问答请求失败 ${row.error_count || 0} 次</li>`);
      const attentionList = itemList(learning.attention, row => `<li><b>${escapeHtml(row.display_name || row.username)}</b> <span class="muted">@${escapeHtml(row.username)}</span><br>${escapeHtml(row.attention_reason)}；最近活动：${escapeHtml(String(row.last_activity || "暂无").slice(0, 19))}</li>`);
      const attentionRules = (learning.attention_rules || []).length
        ? "<ul class='rules'>" + learning.attention_rules.map(rule => `<li>${escapeHtml(rule)}</li>`).join("") + "</ul>"
        : "<p class='muted'>暂无规则</p>";
      const timingRows = (data.response_timings || []).length ? data.response_timings.map(row => {
        const timings = row.timings || {};
        return `<tr>
          <td>${escapeHtml(String(row.timestamp || "").slice(0, 19))}</td>
          <td>${escapeHtml(row.display_name || row.username || "匿名用户")}</td>
          <td>${escapeHtml(String(row.question || "").slice(0, 80))}</td>
          <td>${formatDuration(timings["检索耗时"])}</td>
          <td>${formatDuration(timings["上下文拼装耗时"])}</td>
          <td>${formatDuration(timings["联网检索耗时"])}</td>
          <td>${formatDuration(timings["历史加载耗时"])}</td>
          <td>${formatDuration(timings["模型首片段耗时"])}</td>
          <td>${formatDuration(timings["模型流式总耗时"])}</td>
          <td>${formatDuration(timings["端到端耗时"] ?? row.response_time_ms)}</td>
        </tr>`;
      }).join("") : `<tr><td colspan="10" class="muted">暂无分阶段耗时记录，新问答完成后会自动记录。</td></tr>`;
      const rosterList = itemList(roster.list, row => {
        const typeName = row.identity_type === "student" ? "学生" : "教师";
        const idName = row.identity_type === "student" ? "学号" : "工号";
        const state = row.username ? `已绑定 @${escapeHtml(row.username)}` : "未绑定";
        const label = `${row.real_name} · ${typeName} · ${row.institutional_id}`;
        const actions = row.username ? "" : `<span class="roster-actions"><button type="button" onclick='editRoster(${Number(row.id)}, ${JSON.stringify(row.identity_type)}, ${JSON.stringify(row.institutional_id)}, ${JSON.stringify(row.real_name)})'>修改</button><button type="button" class="danger" onclick='deleteRoster(${Number(row.id)}, ${JSON.stringify(label)})'>删除</button></span>`;
        return `<li><b>${escapeHtml(row.real_name)}</b> · ${typeName} · ${idName} ${escapeHtml(row.institutional_id)}<br><span class="muted">${state}</span>${actions}</li>`;
      });
      dashboard.innerHTML = `
        <div class="grid">
          <div class="card"><div class="muted">总提问数</div><div class="metric">${total.total_questions ?? 0}</div></div>
          <div class="card"><div class="muted">问答请求失败数</div><div class="metric">${total.total_errors ?? 0}</div><div class="muted">失败率 ${total.error_rate ?? "0.0%"}</div></div>
          <div class="card"><div class="muted">用户数</div><div class="metric">${users.total_users ?? 0}</div></div>
          <div class="card"><div class="muted">Sessions</div><div class="metric">${total.total_sessions ?? 0}</div></div>
          <div class="card"><div class="muted">已登录用户</div><div class="metric">${users.active_users ?? 0}</div></div>
          <div class="card"><div class="muted">登录次数</div><div class="metric">${users.total_logins ?? 0}</div></div>
        </div>
        <p class="section-note"><b>统计口径：</b>总览包含注册用户与匿名用户。问答请求失败率 = 最终未能正常完成且记录了异常的问答数 ÷ 总提问数 × 100%。它反映模型接口、网络、视觉识别或服务异常，不判断答案在知识上是否正确；用户差评单独统计，Python 代码与可视化运行错误列入下方“系统运行错误日志”。</p>
        <h2 class="section-title">响应耗时分析</h2>
        <p class="section-note">仅管理员可见，展示最近 30 次新问答的知识库、历史记录和模型生成各阶段耗时。</p>
        <div class="table-wrap">
          <table>
            <thead><tr><th>时间</th><th>用户</th><th>问题</th><th>知识检索</th><th>上下文</th><th>联网检索</th><th>历史加载</th><th>首段答案</th><th>模型总计</th><th>端到端</th></tr></thead>
            <tbody>${timingRows}</tbody>
          </table>
        </div>
        <h2 class="section-title">身份名册</h2>
        <p class="section-note">身份绑定为可选功能。用户填写学生或教师身份时，类型、学号或工号、姓名必须与名册匹配；每个编号只能绑定一个账号。</p>
        <div class="grid">
          <div class="card"><div class="muted">学生名册</div><div class="metric">${roster.student_total ?? 0}</div></div>
          <div class="card"><div class="muted">已绑定学生</div><div class="metric">${roster.student_bound ?? 0}</div></div>
          <div class="card"><div class="muted">教师名册</div><div class="metric">${roster.teacher_total ?? 0}</div></div>
          <div class="card"><div class="muted">已绑定教师</div><div class="metric">${roster.teacher_bound ?? 0}</div></div>
        </div>
        <div class="cols">
          <div class="card">
            <h2>批量录入名册</h2>
            <p class="muted">每行一条：身份类型, 学号或工号, 姓名</p>
            <div class="roster-form">
              <textarea id="roster-input" placeholder="学生, 20260001, 张三&#10;教师, T001, 李老师"></textarea>
              <button onclick="importRoster()">导入名册</button>
            </div>
            <div class="excel-upload">
              <div>
                <input id="roster-excel" type="file" accept=".xlsx,.xlsm,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" />
                <p class="muted">支持 .xlsx/.xlsm；首行包含“身份类型、学号/工号、姓名”，编号建议设为文本格式。</p>
              </div>
              <button id="roster-excel-button" onclick="importRosterExcel()">上传 Excel</button>
            </div>
            <div id="roster-result" class="success"></div>
          </div>
          <div class="card">
            <h2>名册与绑定状态</h2>
            ${rosterList}
          </div>
        </div>
        <h2 class="section-title">学情分析</h2>
        <p class="section-note">仅统计注册学生的实际学习活动。个人及章节“问答请求失败率”均按最终失败问答数 ÷ 对应总提问数计算，只表示服务请求是否完成，不代表答案正确率、课程成绩或知识掌握度。</p>
        <div class="grid">
          <div class="card"><div class="muted">注册学习者</div><div class="metric">${learningOverview.registered_learners ?? 0}</div></div>
          <div class="card"><div class="muted">近 7 日活跃</div><div class="metric">${learningOverview.active_7d ?? 0}</div></div>
          <div class="card"><div class="muted">近 30 日活跃</div><div class="metric">${learningOverview.active_30d ?? 0}</div></div>
          <div class="card"><div class="muted">人均提问</div><div class="metric">${learningOverview.avg_questions_per_learner ?? 0}</div></div>
          <div class="card"><div class="muted">需关注</div><div class="metric">${learningOverview.attention_count ?? 0}</div></div>
        </div>
        <div class="cols">
          <div class="card">
            <h2>章节学习分布</h2>
            ${chapterLearningList}
          </div>
          <div class="card">
            <h2>近 30 日学习趋势</h2>
            ${dailyLearningList}
          </div>
        </div>
        <div class="card" style="margin-top:14px;">
          <h2>需关注的学习者</h2>
          ${attentionList}
          <details><summary>查看判定规则</summary>${attentionRules}</details>
        </div>
        <div style="margin-top:14px;">
          <h2>注册用户学习明细</h2>
          <div class="table-wrap">
            <table>
              <thead><tr><th>用户</th><th>提问</th><th>学习天数</th><th>覆盖章节</th><th>会话</th><th>问答请求失败率</th><th>好评/差评</th><th>最近活动</th><th>状态</th></tr></thead>
              <tbody>${learnerRows}</tbody>
            </table>
          </div>
        </div>
        <div class="cols">
          <div class="card">
            <h2>📚 章节统计</h2>
            ${itemList(data.chapters, r => `<li><b>${escapeHtml(r.chapter || "未分类")}</b>: ${r.cnt} 题 | in=${r.ti || 0} out=${r.to_tokens || 0}</li>`)}
          </div>
          <div class="card">
            <h2>📅 每日用量</h2>
            ${itemList(data.daily, r => `<li><b>${r.day}</b>: ${r.questions} 题 | in=${r.ti || 0} out=${r.to_tokens || 0}</li>`)}
          </div>
        </div>
        <div class="cols" style="margin-top:14px;">
          <div class="card">
            <h2>👥 用户列表</h2>
            ${itemList(users.list, r => `<li><b>${escapeHtml(r.display_name || r.username)}</b> <span class="muted">@${escapeHtml(r.username)}</span><br><span class="muted">${String(r.identity_verified) === "1" ? ((r.identity_type === "student" ? "学生 · 学号 " : "教师 · 工号 ") + escapeHtml(r.institutional_id)) : "未绑定身份"} | 注册：${escapeHtml(String(r.created_at || "").slice(0,19))} | 登录 ${r.login_count || 0} 次 | 最近登录：${r.last_login ? escapeHtml(String(r.last_login).slice(0,19)) : "尚未登录"} | ${String(r.is_active) === "1" ? "正常" : "停用"}</span></li>`)}
          </div>
          <div class="card">
            <h2>📊 登录数据统计</h2>
            ${itemList(users.login_daily, r => `<li><b>${r.day}</b>：${r.logins} 次登录，${r.users} 位用户</li>`)}
          </div>
        </div>
        <div class="card" style="margin-top:14px;">
          <h2>Detailed Feedback</h2>
          ${itemList(data.feedback_recent, r => `<li><b>${escapeHtml(String(r.timestamp || "").slice(0,19))}</b> ${escapeHtml(r.rating)}: ${escapeHtml(r.comment)}<br><span class="muted">${escapeHtml(r.display_name || r.username || "anonymous")}</span></li>`)}
        </div>
        <div class="cols" style="margin-top:14px;">
          <div class="card">
            <h2>👤 最近用户</h2>
            ${itemList(users.recent, r => `<li><b>${escapeHtml(r.display_name || r.username)}</b> <span class="muted">@${escapeHtml(r.username)}</span><br><span class="muted">最近登录: ${r.last_login ? escapeHtml(String(r.last_login).slice(0,19)) : "尚未登录"}</span></li>`)}
          </div>
          <div class="card">
            <h2>💬 反馈统计</h2>
            ${itemList(data.feedback, r => `<li><b>${r.rating}</b>: ${r.cnt} 次</li>`)}
          </div>
        </div>
        <div class="cols" style="margin-top:14px;">
          <div class="card">
            <h2>🚨 系统运行错误日志</h2>
            ${itemList((data.errors || {}).recent, r => `<li><b>${escapeHtml(String(r.timestamp).slice(0,19))}</b> ${escapeHtml(r.error_type)}<br><span class="muted">${escapeHtml(r.display_name || r.username || "匿名用户")} · ${escapeHtml(r.question || "未记录问题")}</span><br>${escapeHtml(r.error_message)}<details><summary>查看详细日志</summary><pre>${escapeHtml(r.traceback || "无详细堆栈")}</pre></details></li>`)}
          </div>
          <div class="card">
            <h2>❓ 未成功回答</h2>
            ${itemList(data.unanswered, r => `<li><b>${escapeHtml(r.chapter || "")}</b> ${escapeHtml(String(r.question || "").slice(0,120))}<br><span class="muted">${escapeHtml(String(r.error || "").slice(0,160))}</span></li>`)}
          </div>
        </div>
      `;
    }
    __AUTO_LOAD__
  </script>
</body>
</html>
"""
    auto_script = "window.setTimeout(loadAnalytics, 0);" if auto_load else ""
    auth_note = (
        "管理员身份已验证，正在加载用户、问答、反馈和学情信息。"
        if auto_load else
        "请先从学生端登录管理员账号；也可以输入管理员访问令牌。"
    )
    auth_style = 'style="display:none"' if auto_load else ""
    prefix_json = json.dumps(public_prefix.rstrip("/"))
    page = page.replace("const tokenInput = document.getElementById(\"token\");", f"const API_PREFIX = {prefix_json};\n    const apiUrl = path => API_PREFIX + path;\n\n    const tokenInput = document.getElementById(\"token\");")
    page = page.replace('fetch("/identity-roster",', 'fetch(apiUrl("/identity-roster"),')
    page = page.replace('fetch("/identity-roster/excel",', 'fetch(apiUrl("/identity-roster/excel"),')
    page = page.replace('fetch("/analytics?format=json",', 'fetch(apiUrl("/analytics?format=json"),')
    return (page.replace("__AUTO_LOAD__", auto_script)
                .replace("__AUTH_NOTE__", auth_note)
                .replace("__AUTH_STYLE__", auth_style))


app = FastAPI(
    title="大学物理智能助教 Admin API",
    description="管理员数据分析接口。需要 X-Admin-Token 或 Bearer token。",
    version="1.0.0",
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[],
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["Authorization", "X-Admin-Token"],
)


@app.get("/health")
def health():
    return {"ok": True}


def _request_public_prefix(request: Request) -> str:
    public_prefix = request.headers.get("x-forwarded-prefix", "").strip()
    if not public_prefix:
        public_prefix = os.getenv("PHYSICS_GATEWAY_PUBLIC_PREFIX", "")
    normalized = "/" + public_prefix.strip("/")
    return "" if normalized == "/" else normalized


def _public_app_url(request: Request, mode: str) -> str:
    safe_mode = mode if mode in {"light", "system", "dark"} else "system"
    return with_public_prefix(f"/?mode={safe_mode}", "", _request_public_prefix(request))


def _user_session_cookie_path(request: Request) -> str:
    return _request_public_prefix(request) or "/"


def _request_is_https(request: Request) -> bool:
    forwarded_proto = request.headers.get(
        "x-forwarded-proto", request.url.scheme
    ).split(",", 1)[0].strip()
    return forwarded_proto == "https"


@app.get("/session-login")
def user_login_session(
    request: Request,
    ticket: str = Query(min_length=20, max_length=4096),
    mode: str = Query(default="system", max_length=16),
):
    """Exchange a one-minute login ticket for an HttpOnly browser session."""
    secret = _load_admin_token()
    payload = user_session.verify_login_ticket(secret, ticket)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired user login ticket.")

    now = int(time.time())
    for nonce, expiry in list(_USED_USER_LOGIN_NONCES.items()):
        if expiry < now:
            _USED_USER_LOGIN_NONCES.pop(nonce, None)
    nonce = str(payload["nonce"])

    username = str(payload["sub"])
    account = db.get_user_by_username(username)
    if not account or not account.get("is_active"):
        raise HTTPException(status_code=403, detail="User account is not active.")
    if nonce in _USED_USER_LOGIN_NONCES:
        existing_account = user_session.resolve_session(
            secret,
            request.cookies.get(user_session.USER_SESSION_COOKIE, ""),
            db.get_user_by_username,
        )
        if (
            not existing_account
            or str(existing_account.get("username", "")).casefold()
            != username.casefold()
        ):
            raise HTTPException(
                status_code=401,
                detail="User login ticket has already been used.",
            )
        return RedirectResponse(url=_public_app_url(request, mode), status_code=303)

    _USED_USER_LOGIN_NONCES[nonce] = int(payload["exp"])

    session_seconds = user_session.configured_session_seconds(
        os.getenv("PHYSICS_USER_SESSION_SECONDS")
    )
    response = RedirectResponse(url=_public_app_url(request, mode), status_code=303)
    response.set_cookie(
        user_session.USER_SESSION_COOKIE,
        user_session.issue_session(secret, username, session_seconds),
        max_age=session_seconds,
        httponly=True,
        secure=_request_is_https(request),
        samesite="strict",
        path=_user_session_cookie_path(request),
    )
    return response


@app.get("/session-logout")
def user_logout_session(
    request: Request,
    ticket: str = Query(min_length=20, max_length=4096),
    mode: str = Query(default="system", max_length=16),
):
    """Clear the persistent login cookie after validating a signed logout request."""
    if not user_session.verify_logout_ticket(_load_admin_token(), ticket):
        raise HTTPException(status_code=401, detail="Invalid or expired user logout ticket.")
    response = RedirectResponse(url=_public_app_url(request, mode), status_code=303)
    response.delete_cookie(
        user_session.USER_SESSION_COOKIE,
        path=_user_session_cookie_path(request),
        secure=_request_is_https(request),
        httponly=True,
        samesite="strict",
    )
    return response


@app.get("/admin-login")
def admin_login(request: Request, ticket: str = Query(min_length=20, max_length=4096)):
    secret = _load_admin_token()
    payload = admin_auth.verify_token(secret, ticket, "admin-login")
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired administrator login ticket.")

    now = int(time.time())
    for nonce, expiry in list(_USED_LOGIN_NONCES.items()):
        if expiry < now:
            _USED_LOGIN_NONCES.pop(nonce, None)
    nonce = str(payload["nonce"])
    if nonce in _USED_LOGIN_NONCES:
        raise HTTPException(status_code=401, detail="Administrator login ticket has already been used.")

    user = db.get_user_by_username(str(payload["sub"]))
    if not user or user.get("role") != "admin" or not user.get("is_active"):
        raise HTTPException(status_code=403, detail="Administrator account is not active.")
    _USED_LOGIN_NONCES[nonce] = int(payload["exp"])

    session_token = admin_auth.issue_token(
        secret, str(payload["sub"]), "admin-session", _ADMIN_SESSION_SECONDS
    )
    # The gateway may mount this API below /agent (or another prefix).  Keep
    # the browser on that public path after ticket authentication.
    public_prefix = request.headers.get("x-forwarded-prefix", "").strip()
    if not public_prefix:
        public_prefix = os.getenv("PHYSICS_GATEWAY_PUBLIC_PREFIX", "")
    response = RedirectResponse(
        url=with_public_prefix("/analytics", "", public_prefix), status_code=303
    )
    forwarded_proto = request.headers.get("x-forwarded-proto", request.url.scheme).split(",", 1)[0].strip()
    response.set_cookie(
        _ADMIN_SESSION_COOKIE,
        session_token,
        max_age=_ADMIN_SESSION_SECONDS,
        httponly=True,
        secure=forwarded_proto == "https",
        samesite="strict",
        path="/",
    )
    return response


@app.post("/identity-roster")
def update_identity_roster(
    payload: IdentityRosterBatch,
    request: Request,
    x_admin_token: str | None = Header(default=None),
    authorization: str | None = Header(default=None),
    admin_session: str | None = Cookie(default=None, alias=_ADMIN_SESSION_COOKIE),
):
    client_ip = request.client.host if request.client else "unknown"
    _require_admin_token(x_admin_token, authorization, client_ip, admin_session)
    entries = [entry.model_dump() for entry in payload.entries]
    return db.upsert_identity_roster(entries)


@app.put("/identity-roster/{roster_id}")
def edit_identity_roster(
    roster_id: int,
    payload: IdentityRosterEdit,
    request: Request,
    x_admin_token: str | None = Header(default=None),
    authorization: str | None = Header(default=None),
    admin_session: str | None = Cookie(default=None, alias=_ADMIN_SESSION_COOKIE),
):
    client_ip = request.client.host if request.client else "unknown"
    _require_admin_token(x_admin_token, authorization, client_ip, admin_session)
    try:
        return db.update_identity_roster_entry(roster_id, payload.identity_type, payload.institutional_id, payload.real_name)
    except (ValueError, LookupError, PermissionError) as exc:
        raise HTTPException(status_code=409 if isinstance(exc, PermissionError) else 400, detail=str(exc)) from exc


@app.delete("/identity-roster/{roster_id}")
def remove_identity_roster(
    roster_id: int,
    request: Request,
    x_admin_token: str | None = Header(default=None),
    authorization: str | None = Header(default=None),
    admin_session: str | None = Cookie(default=None, alias=_ADMIN_SESSION_COOKIE),
):
    client_ip = request.client.host if request.client else "unknown"
    _require_admin_token(x_admin_token, authorization, client_ip, admin_session)
    try:
        return db.delete_identity_roster_entry(roster_id)
    except (LookupError, PermissionError) as exc:
        raise HTTPException(status_code=409 if isinstance(exc, PermissionError) else 404, detail=str(exc)) from exc


@app.post("/identity-roster/excel")
async def upload_identity_roster_excel(
    request: Request,
    x_admin_token: str | None = Header(default=None),
    authorization: str | None = Header(default=None),
    admin_session: str | None = Cookie(default=None, alias=_ADMIN_SESSION_COOKIE),
):
    client_ip = request.client.host if request.client else "unknown"
    _require_admin_token(x_admin_token, authorization, client_ip, admin_session)
    content = await request.body()
    try:
        entries, parse_errors = _parse_identity_roster_excel(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not entries:
        detail = "Excel 中没有可导入的有效记录。"
        if parse_errors:
            detail += " " + "；".join(parse_errors[:10])
        raise HTTPException(status_code=400, detail=detail)

    result = db.upsert_identity_roster(entries)
    result["rows_read"] = len(entries)
    result["errors"] = parse_errors + list(result.get("errors", []))
    return result


@app.get("/analytics")
def analytics(
    request: Request,
    x_admin_token: str | None = Header(default=None),
    authorization: str | None = Header(default=None),
    admin_session: str | None = Cookie(default=None, alias=_ADMIN_SESSION_COOKIE),
    recent_error_limit: int = Query(default=15, ge=1, le=100),
    format: str = Query(default="html"),
):
    public_prefix = request.headers.get("x-forwarded-prefix", "").strip()
    if not public_prefix:
        public_prefix = os.getenv("PHYSICS_GATEWAY_PUBLIC_PREFIX", "").strip()
    public_prefix = "/" + public_prefix.strip("/") if public_prefix.strip("/") else ""
    has_session = _valid_admin_session(admin_session)
    has_token = bool(x_admin_token or authorization)
    if not has_session and not has_token and format != "json":
        return HTMLResponse(_analytics_login_page(public_prefix=public_prefix))

    client_ip = request.client.host if request.client else "unknown"
    _require_admin_token(x_admin_token, authorization, client_ip, admin_session)
    if format != "json":
        return HTMLResponse(_analytics_login_page(auto_load=True, public_prefix=public_prefix))
    return _analytics_payload(recent_error_limit)
